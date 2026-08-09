import io, re, json, hashlib, difflib
from copy import copy
from collections import Counter
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation

import numpy as np
import pandas as pd
import pymupdf
import pytesseract
import streamlit as st
from PIL import Image, ImageOps, ImageSequence, ImageFilter
from openpyxl import load_workbook


# ============================================================
# CONFIG
# ============================================================

st.set_page_config(
    page_title="Vendor Invoice Processor",
    page_icon="🧾",
    layout="wide",
    initial_sidebar_state="collapsed",
)

PACK_TYPES = {
    "Cigarettes (10)": 10,
    "Dip / Smokeless (5)": 5,
    "Each (1)": 1,
}

RECEIPT_COLUMNS = [
    "Row Type",
    "Item Number",
    "Description",
    "Qty Purchased",
    "Qty Returned",
    "Purchase Total",
    "Credit / Adjustment",
    "Net Amount",
    "Pack Type",
    "Cost Variance?",
]

DEFAULT_STATE = {
    "theme_mode": "Dark",
    "processed": False,
    "ocr_mode": "Standard",
    "clover_bytes": None,
    "clover_df": None,
    "clover_header_index": None,
    "receipt_rows": None,
    "raw_ocr": "",
    "prepared_output": None,
    "prepared_filename": None,
    "prepared_fingerprint": None,
    "prepared_diagnostics": None,
}

for k, v in DEFAULT_STATE.items():
    if k not in st.session_state:
        st.session_state[k] = v


# ============================================================
# HELPERS
# ============================================================

def text_value(v):
    if v is None or (
        isinstance(v, float)
        and np.isnan(v)
    ):
        return ""

    return str(v).strip()


def normalize_name(v):
    return re.sub(
        r"\s+",
        " ",
        re.sub(
            r"[^A-Z0-9 ]+",
            " ",
            text_value(v).upper(),
        ),
    ).strip()


def normalize_digits(v):
    s = text_value(v)

    if s.endswith(".0"):
        s = s[:-2]

    return re.sub(
        r"[^0-9]",
        "",
        s,
    )


def to_number(v):
    if v is None:
        return None

    if isinstance(
        v,
        (
            int,
            float,
            np.integer,
            np.floating,
        ),
    ):
        return (
            None
            if pd.isna(v)
            else float(v)
        )

    s = (
        str(v)
        .replace("$", "")
        .replace(",", "")
        .strip()
    )

    if (
        not s
        or s.lower() == "nan"
    ):
        return None

    try:
        return float(s)

    except ValueError:
        return None


def dec(v):
    n = to_number(v)

    if n is None:
        return None

    try:
        return Decimal(str(n))

    except (
        InvalidOperation,
        TypeError,
        ValueError,
    ):
        return None


def money_round(v):
    d = dec(v)

    if d is None:
        return None

    return float(
        d.quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )
    )


def divide_currency(
    total,
    qty,
):
    t = dec(total)
    q = dec(qty)

    if (
        t is None
        or q is None
        or q <= 0
    ):
        return None

    return float(
        (t / q).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )
    )


def margin_percent(
    cost,
    retail,
):
    c = to_number(cost)
    r = to_number(retail)

    # $0 Clover cost = missing/unknown.
    if (
        c is None
        or c <= 0
        or r is None
        or r <= 0
    ):
        return np.nan

    return round(
        ((r - c) / r) * 100,
        1,
    )


def retail_for_margin(
    cost,
    margin,
):
    c = to_number(cost)
    m = to_number(margin)

    if (
        c is None
        or c <= 0
        or m is None
        or m < 0
        or m >= 100
    ):
        return None

    denominator = (
        Decimal("1")
        - Decimal(str(m))
        / Decimal("100")
    )

    if denominator <= 0:
        return None

    return float(
        (
            Decimal(str(c))
            / denominator
        ).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )
    )


def values_different(
    a,
    b,
    tolerance=0.004,
):
    x = to_number(a)
    y = to_number(b)

    if (
        x is not None
        and y is not None
    ):
        return (
            abs(x - y)
            > tolerance
        )

    return (
        text_value(a)
        != text_value(b)
    )


def clear_prepared_output():
    for k in (
        "prepared_output",
        "prepared_filename",
        "prepared_fingerprint",
        "prepared_diagnostics",
    ):
        st.session_state[k] = None


def clear_mapping_state():
    prefixes = (
        "action_",
        "match_",
        "search_",
        "cost_action_",
        "retail_",
        "variance_confirm_",
        "new_name_",
        "new_upc_",
        "new_retail_",
        "new_category_",
        "new_tax_",
        "new_printer_",
    )

    for k in list(
        st.session_state.keys()
    ):
        if k.startswith(prefixes):
            st.session_state.pop(
                k,
                None,
            )


def reset_processed_data():
    st.session_state.processed = False
    st.session_state.ocr_mode = "Standard"
    st.session_state.clover_bytes = None
    st.session_state.clover_df = None
    st.session_state.clover_header_index = None
    st.session_state.receipt_rows = None
    st.session_state.raw_ocr = ""

    clear_mapping_state()
    clear_prepared_output()

    st.session_state.pop(
        "receipt_editor",
        None,
    )


# ============================================================
# THEME
# ============================================================

def apply_theme(theme):

    if theme == "Light":

        bg = "#f4f7fb"
        panel = "#ffffff"
        panel2 = "#f9fafb"
        text = "#101828"
        muted = "#667085"
        border = "#d9e0e8"
        accent = "#1677ff"
        accent2 = "#12b76a"

    else:

        bg = "#0b0f16"
        panel = "#101722"
        panel2 = "#151d2a"
        text = "#f7f9fc"
        muted = "#98a2b3"
        border = "#344054"
        accent = "#4da3ff"
        accent2 = "#32d583"

    st.markdown(
        f"""
        <style>

        .stApp {{
            background: {bg};
            color: {text};
        }}

        .block-container {{
            max-width: 1500px;
            padding-top: 1.5rem;
            padding-bottom: 5rem;
        }}

        .stApp h1,
        .stApp h2,
        .stApp h3,
        .stApp h4,
        .stApp p,
        .stApp label {{
            color: {text};
        }}

        div[data-testid="stCaptionContainer"] p {{
            color: {muted} !important;
        }}

        div[data-testid="stVerticalBlockBorderWrapper"] {{
            background: {panel};
            border-color: {border};
            border-radius: 18px;
            box-shadow: 0 8px 28px rgba(0,0,0,.08);
        }}

        section[data-testid="stFileUploaderDropzone"] {{
            background: {panel2} !important;
            border-color: {border} !important;
            border-radius: 14px !important;
        }}

        div[data-testid="stMetric"] {{
            background: {panel2};
            border: 1px solid {border};
            border-radius: 14px;
            padding: 14px;
        }}

        div[data-testid="stButton"] button {{
            border-radius: 12px;
            min-height: 46px;
        }}

        div[data-testid="stButton"] button[kind="primary"] {{
            background: linear-gradient(
                90deg,
                {accent},
                {accent2}
            ) !important;

            border: 0 !important;
            color: white !important;
            min-height: 52px;
            font-weight: 750;
        }}

        div[data-testid="stButton"]
        button[kind="primary"] * {{
            color: white !important;
        }}

        div[data-testid="stDownloadButton"] button {{
            background: linear-gradient(
                90deg,
                {accent2},
                {accent}
            ) !important;

            border: 0 !important;
            color: white !important;
            min-height: 58px;
            border-radius: 12px;
            font-weight: 800;
        }}

        div[data-testid="stDownloadButton"] button * {{
            color: white !important;
        }}

        button[data-baseweb="tab"] {{
            font-weight: 750;
        }}

        #MainMenu {{
            visibility: hidden;
        }}

        footer {{
            visibility: hidden;
        }}

        </style>
        """,
        unsafe_allow_html=True,
    )


# ============================================================
# CLOVER
# ============================================================

def detect_clover_header_index(
    clover_bytes,
):

    preview = pd.read_excel(
        io.BytesIO(
            clover_bytes
        ),
        sheet_name="Items",
        header=None,
        nrows=15,
        dtype=object,
    )

    required = {
        "Clover ID",
        "Name",
        "Price",
        "Cost",
    }

    for i, row in (
        preview.iterrows()
    ):

        values = {
            text_value(v)
            for v
            in row.tolist()
        }

        if required.issubset(
            values
        ):
            return int(i)

    raise ValueError(
        "Could not locate the Clover Items header row."
    )


def load_clover(
    clover_bytes,
):

    header_index = (
        detect_clover_header_index(
            clover_bytes
        )
    )

    raw = pd.read_excel(
        io.BytesIO(
            clover_bytes
        ),
        sheet_name="Items",
        header=header_index,
        dtype=object,
    )

    raw.columns = [
        text_value(c)
        for c
        in raw.columns
    ]

    missing = [
        c
        for c in (
            "Clover ID",
            "Name",
            "Price",
            "Cost",
        )
        if c not in raw.columns
    ]

    if missing:
        raise ValueError(
            "Clover Items sheet is missing: "
            + ", ".join(
                missing
            )
        )

    df = raw.loc[
        raw["Name"].notna()
        | raw["Clover ID"].notna()
    ].copy()

    df["_Excel Row"] = (
        df.index
        + header_index
        + 2
    )

    df["_Name Clean"] = (
        df["Name"]
        .apply(
            normalize_name
        )
    )

    if (
        "Product Code"
        in df.columns
    ):
        df[
            "_Product Code Clean"
        ] = (
            df[
                "Product Code"
            ]
            .apply(
                normalize_digits
            )
        )

    else:
        df[
            "_Product Code Clean"
        ] = ""

    if "SKU" in df.columns:

        df["_SKU Clean"] = (
            df["SKU"]
            .apply(
                normalize_name
            )
        )

    else:
        df["_SKU Clean"] = ""

    df = df.reset_index(
        drop=True
    )

    # This exists only during the current run.
    # Nothing is permanently remembered.
    df["_Match Key"] = [
        (
            f"ID::{text_value(cid)}"
            if text_value(cid)
            else f"ROW::{i}"
        )
        for i, cid
        in enumerate(
            df[
                "Clover ID"
            ].tolist()
        )
    ]

    return (
        df,
        header_index,
    )


def clover_categories(df):

    values = {
        "Tobacco"
    }

    if (
        "Categories"
        in df.columns
    ):

        values |= {
            str(v).strip()
            for v
            in df[
                "Categories"
            ].dropna().tolist()
            if str(v).strip()
        }

    return sorted(
        values
    )


# ============================================================
# OCR
# ============================================================

def rotate_image(
    image,
    rotation,
):

    if rotation == 90:
        return image.rotate(
            -90,
            expand=True,
        )

    if rotation == 180:
        return image.rotate(
            180,
            expand=True,
        )

    if rotation == 270:
        return image.rotate(
            -270,
            expand=True,
        )

    return image


def prepare_image(
    image,
    mode="Standard",
    rotation=0,
):

    image = rotate_image(
        ImageOps.exif_transpose(
            image
        ),
        rotation,
    ).convert(
        "L"
    )

    image = ImageOps.autocontrast(
        image
    )

    target_width = (
        1800
        if mode == "Standard"
        else 2200
    )

    w, h = image.size

    if (
        w > 0
        and w != target_width
    ):

        scale = (
            target_width
            / w
        )

        image = image.resize(
            (
                target_width,
                max(
                    1,
                    int(
                        h * scale
                    ),
                ),
            ),
            Image.Resampling.LANCZOS,
        )

    if mode == "Strong":

        image = image.filter(
            ImageFilter.UnsharpMask(
                radius=2,
                percent=170,
                threshold=3,
            )
        )

        image = image.point(
            lambda p:
            255
            if p > 178
            else 0
        )

    return image


def merge_ocr_blocks(
    blocks,
):
    """
    Removes overlap created by chunking long images.

    A single repeated receipt line is deliberately
    NOT removed. At least two consecutive matching
    lines are required before an overlap is discarded.
    """

    combined = []

    for block in blocks:

        lines = [
            re.sub(
                r"\s+",
                " ",
                x,
            ).strip()

            for x
            in block.splitlines()

            if re.sub(
                r"\s+",
                " ",
                x,
            ).strip()
        ]

        if not lines:
            continue

        if not combined:
            combined.extend(
                lines
            )
            continue

        drop = 0

        for overlap in range(
            min(
                12,
                len(combined),
                len(lines),
            ),
            1,
            -1,
        ):

            left = [
                normalize_name(v)
                for v
                in combined[
                    -overlap:
                ]
            ]

            right = [
                normalize_name(v)
                for v
                in lines[
                    :overlap
                ]
            ]

            if (
                left == right
                and all(left)
            ):
                drop = overlap
                break

        combined.extend(
            lines[
                drop:
            ]
        )

    return "\n".join(
        combined
    )


def ocr_prepared_image(
    image,
    mode="Standard",
):

    if mode == "Standard":

        max_chunk = 3600
        overlap = 80
        config = (
            "--oem 1 "
            "--psm 6"
        )

    else:

        max_chunk = 2800
        overlap = 140
        config = (
            "--oem 1 "
            "--psm 11"
        )

    if (
        image.height
        <= max_chunk
    ):

        return (
            pytesseract
            .image_to_string(
                image,
                lang="eng",
                config=config,
            )
        )

    blocks = []
    start = 0

    while (
        start
        < image.height
    ):

        end = min(
            image.height,
            start
            + max_chunk,
        )

        crop = image.crop(
            (
                0,
                start,
                image.width,
                end,
            )
        )

        blocks.append(
            pytesseract
            .image_to_string(
                crop,
                lang="eng",
                config=config,
            )
        )

        if (
            end
            >= image.height
        ):
            break

        start = max(
            0,
            end
            - overlap,
        )

    return merge_ocr_blocks(
        blocks
    )


def read_single_file(
    uploaded_file,
    mode="Standard",
    rotation=0,
):

    data = (
        uploaded_file
        .getvalue()
    )

    name = (
        uploaded_file
        .name
        .lower()
    )

    pages = []

    # PDF / scanned PDF
    if name.endswith(
        ".pdf"
    ):

        doc = pymupdf.open(
            stream=data,
            filetype="pdf",
        )

        try:

            for n, page in enumerate(
                doc,
                1,
            ):

                embedded = (
                    page
                    .get_text(
                        "text"
                    )
                    .strip()
                )

                # If the PDF already has good text,
                # use it instead of OCR in Standard mode.
                if (
                    mode == "Standard"
                    and len(
                        re.sub(
                            r"\s+",
                            "",
                            embedded,
                        )
                    )
                    >= 40
                ):

                    text = embedded

                else:

                    zoom = (
                        1.7
                        if mode == "Standard"
                        else 2.1
                    )

                    pix = (
                        page.get_pixmap(
                            matrix=(
                                pymupdf.Matrix(
                                    zoom,
                                    zoom,
                                )
                            ),
                            alpha=False,
                        )
                    )

                    img = (
                        Image.frombytes(
                            "RGB",
                            (
                                pix.width,
                                pix.height,
                            ),
                            pix.samples,
                        )
                    )

                    text = (
                        ocr_prepared_image(
                            prepare_image(
                                img,
                                mode,
                                rotation,
                            ),
                            mode,
                        )
                    )

                pages.append(
                    f"--- "
                    f"{uploaded_file.name} "
                    f"/ PAGE {n} ---\n"
                    f"{text}"
                )

        finally:
            doc.close()

        return "\n".join(
            pages
        )

    # JPG / PNG / TIFF
    image = Image.open(
        io.BytesIO(
            data
        )
    )

    for n, frame in enumerate(
        ImageSequence.Iterator(
            image
        ),
        1,
    ):

        text = (
            ocr_prepared_image(
                prepare_image(
                    frame.copy(),
                    mode,
                    rotation,
                ),
                mode,
            )
        )

        pages.append(
            f"--- "
            f"{uploaded_file.name} "
            f"/ IMAGE {n} ---\n"
            f"{text}"
        )

    return "\n".join(
        pages
    )


def read_receipt_files(
    files,
    mode="Standard",
    rotation=0,
):

    return "\n".join(
        read_single_file(
            f,
            mode,
            rotation,
        )
        for f
        in files
    )


# ============================================================
# COSTCO PARSER
# ============================================================

MONEY_RE = re.compile(
    r"(?<!\d)"
    r"(\(?-?\$?\s*"
    r"\d{1,6}[.,]\d{2}"
    r"-?\)?)"
    r"(?!\d)"
)


def parse_money_token(
    token,
):

    raw = (
        text_value(
            token
        )
        .replace(
            "$",
            "",
        )
        .replace(
            " ",
            "",
        )
    )

    negative = False

    if (
        raw.startswith("(")
        and raw.endswith(")")
    ):
        negative = True
        raw = raw[
            1:-1
        ]

    if raw.startswith("-"):
        negative = True
        raw = raw[
            1:
        ]

    if raw.endswith("-"):
        negative = True
        raw = raw[
            :-1
        ]

    try:
        value = float(
            raw.replace(
                ",",
                ".",
            )
        )

    except ValueError:
        return None

    return (
        -value
        if negative
        else value
    )


def choose_description(
    values,
):

    values = [
        text_value(v)
        for v
        in values
        if text_value(v)
    ]

    if not values:
        return "UNKNOWN"

    counts = Counter(
        values
    )

    return max(
        values,
        key=lambda v: (
            counts[v],
            len(v),
        ),
    )


def parse_costco_receipt(
    text,
):

    blocked = [
        "TOTAL NUMBER OF ITEMS",
        "RESALE TOTAL",
        "NON RESALE TOTAL",
        "SUBTOTAL",
        "TOTAL",
        "TAX",
        "CHANGE",
        "TENDER",
        "APPROVED",
        "MEMBER",
        "BOTTOM OF BASKET",
        "BOB COUNT",
        "APP#",
        "TRAN ID",
    ]

    occurrences = []

    for raw_line in (
        text.splitlines()
    ):

        line = re.sub(
            r"\s+",
            " ",
            raw_line,
        ).strip()

        if (
            not line
            or any(
                word
                in line.upper()
                for word
                in blocked
            )
        ):
            continue

        item_match = re.search(
            r"(?<!\d)"
            r"(\d{5,12})"
            r"(?!\d)",
            line,
        )

        amounts = list(
            MONEY_RE.finditer(
                line
            )
        )

        if (
            item_match is None
            or not amounts
        ):
            continue

        amount_match = (
            amounts[
                -1
            ]
        )

        amount = (
            parse_money_token(
                amount_match
                .group(1)
            )
        )

        if amount is None:
            continue

        occurrences.append(
            {
                "Item Number":
                    item_match
                    .group(1),

                "Description":
                    (
                        line[
                            item_match.end()
                            :
                            amount_match.start()
                        ]
                        .strip(
                            " -:$"
                        )
                        or "UNKNOWN"
                    ),

                "Amount":
                    money_round(
                        amount
                    ),
            }
        )

    if not occurrences:

        return pd.DataFrame(
            columns=(
                RECEIPT_COLUMNS
            )
        )

    occurrence_df = (
        pd.DataFrame(
            occurrences
        )
    )

    rows = []

    # Same Costco item number becomes ONE row,
    # regardless of how many separate receipt
    # lines contain it.
    for (
        item_number,
        group,
    ) in occurrence_df.groupby(
        "Item Number",
        sort=False,
    ):

        positive = [
            to_number(v)
            for v
            in group[
                "Amount"
            ].tolist()
            if (
                to_number(v)
                is not None
                and to_number(v)
                > 0
            )
        ]

        negative = [
            to_number(v)
            for v
            in group[
                "Amount"
            ].tolist()
            if (
                to_number(v)
                is not None
                and to_number(v)
                < 0
            )
        ]

        purchase_total = (
            money_round(
                sum(
                    positive
                )
            )
            or 0.0
        )

        credit_total = (
            money_round(
                sum(
                    negative
                )
            )
            or 0.0
        )

        rows.append(
            {
                "Row Type":
                    "Item",

                "Item Number":
                    item_number,

                "Description":
                    choose_description(
                        group[
                            "Description"
                        ].tolist()
                    ),

                "Qty Purchased":
                    len(
                        positive
                    ),

                "Qty Returned":
                    len(
                        negative
                    ),

                "Purchase Total":
                    purchase_total,

                "Credit / Adjustment":
                    credit_total,

                "Net Amount":
                    (
                        money_round(
                            purchase_total
                            + credit_total
                        )
                        or 0.0
                    ),

                "Pack Type":
                    "Cigarettes (10)",

                "Cost Variance?":
                    (
                        "YES - review"
                        if len(
                            {
                                money_round(v)
                                for v
                                in positive
                            }
                        )
                        > 1
                        else ""
                    ),
            }
        )

    return pd.DataFrame(
        rows,
        columns=RECEIPT_COLUMNS,
    )


# ============================================================
# RECEIPT REVIEW / RECONCILIATION
# ============================================================

def ensure_receipt_columns(
    df,
):

    result = df.copy()

    for c in RECEIPT_COLUMNS:

        if c not in result.columns:
            result[c] = None

    return result[
        RECEIPT_COLUMNS
    ]


def recompute_receipt_rows(
    df,
):

    result = (
        ensure_receipt_columns(
            df
        )
    )

    for i, row in (
        result.iterrows()
    ):

        row_type = (
            text_value(
                row.get(
                    "Row Type"
                )
            )
            or "Item"
        )

        if (
            row_type
            == "Adjustment"
        ):

            adjustment = (
                to_number(
                    row.get(
                        "Credit / Adjustment"
                    )
                )
                or 0.0
            )

            result.at[
                i,
                "Qty Purchased",
            ] = 0

            result.at[
                i,
                "Qty Returned",
            ] = 0

            result.at[
                i,
                "Purchase Total",
            ] = 0.0

            result.at[
                i,
                "Credit / Adjustment",
            ] = money_round(
                adjustment
            )

            result.at[
                i,
                "Net Amount",
            ] = money_round(
                adjustment
            )

            continue

        q1 = int(
            to_number(
                row.get(
                    "Qty Purchased"
                )
            )
            or 0
        )

        q2 = int(
            to_number(
                row.get(
                    "Qty Returned"
                )
            )
            or 0
        )

        purchase = (
            to_number(
                row.get(
                    "Purchase Total"
                )
            )
            or 0.0
        )

        credit = (
            to_number(
                row.get(
                    "Credit / Adjustment"
                )
            )
            or 0.0
        )

        result.at[
            i,
            "Row Type",
        ] = "Item"

        result.at[
            i,
            "Qty Purchased",
        ] = max(
            0,
            q1,
        )

        result.at[
            i,
            "Qty Returned",
        ] = max(
            0,
            q2,
        )

        result.at[
            i,
            "Purchase Total",
        ] = money_round(
            max(
                0.0,
                purchase,
            )
        )

        result.at[
            i,
            "Credit / Adjustment",
        ] = money_round(
            credit
        )

        result.at[
            i,
            "Net Amount",
        ] = money_round(
            max(
                0.0,
                purchase,
            )
            + credit
        )

        if not text_value(
            row.get(
                "Pack Type"
            )
        ):

            result.at[
                i,
                "Pack Type",
            ] = (
                "Cigarettes (10)"
            )

    return result


def validate_receipt_rows(
    df,
):

    if (
        df is None
        or df.empty
    ):
        return [
            "No receipt lines are available."
        ]

    errors = []

    for i, row in (
        df.reset_index(
            drop=True
        ).iterrows()
    ):

        display = i + 1

        row_type = (
            text_value(
                row.get(
                    "Row Type"
                )
            )
            or "Item"
        )

        if (
            row_type
            == "Adjustment"
        ):

            amount = (
                to_number(
                    row.get(
                        "Credit / Adjustment"
                    )
                )
                or 0.0
            )

            if (
                abs(amount)
                <= 0.0001
            ):

                errors.append(
                    f"Receipt row "
                    f"{display}: "
                    "Adjustment row "
                    "has no amount."
                )

            continue

        item_number = (
            normalize_digits(
                row.get(
                    "Item Number"
                )
            )
        )

        q1 = int(
            to_number(
                row.get(
                    "Qty Purchased"
                )
            )
            or 0
        )

        q2 = int(
            to_number(
                row.get(
                    "Qty Returned"
                )
            )
            or 0
        )

        purchase = (
            to_number(
                row.get(
                    "Purchase Total"
                )
            )
            or 0.0
        )

        credit = (
            to_number(
                row.get(
                    "Credit / Adjustment"
                )
            )
            or 0.0
        )

        pack_type = (
            text_value(
                row.get(
                    "Pack Type"
                )
            )
        )

        if not item_number:

            errors.append(
                f"Receipt row "
                f"{display}: "
                "Item row is missing "
                "the Costco item number."
            )

        if (
            q1 > 0
            and purchase <= 0
        ):

            errors.append(
                f"Receipt row "
                f"{display}: "
                "purchased quantity "
                "exists but Purchase "
                "Total is zero."
            )

        if (
            purchase > 0
            and q1 <= 0
        ):

            errors.append(
                f"Receipt row "
                f"{display}: "
                "Purchase Total exists "
                "but Qty Bought is zero."
            )

        if (
            q2 > 0
            and credit >= 0
        ):

            errors.append(
                f"Receipt row "
                f"{display}: "
                "Qty Returned exists but "
                "the return/credit amount "
                "is not negative."
            )

        if (
            pack_type
            not in PACK_TYPES
        ):

            errors.append(
                f"Receipt row "
                f"{display}: "
                "select a valid Pack Type."
            )

    return errors


def calculated_receipt_total(
    df,
):

    if (
        df is None
        or df.empty
    ):
        return 0.0

    value = (
        pd.to_numeric(
            df[
                "Net Amount"
            ],
            errors="coerce",
        )
        .fillna(0)
        .sum()
    )

    return (
        money_round(
            value
        )
        or 0.0
    )


def calculated_item_count(
    df,
):

    if (
        df is None
        or df.empty
    ):
        return 0

    rows = df[
        df[
            "Row Type"
        ]
        .astype(str)
        .eq(
            "Item"
        )
    ]

    return int(
        pd.to_numeric(
            rows[
                "Qty Purchased"
            ],
            errors="coerce",
        )
        .fillna(0)
        .sum()
    )


def build_product_groups(
    df,
):

    if (
        df is None
        or df.empty
    ):
        return pd.DataFrame()

    rows = df[
        df[
            "Row Type"
        ]
        .astype(str)
        .eq(
            "Item"
        )
    ].copy()

    rows[
        "Item Number"
    ] = (
        rows[
            "Item Number"
        ]
        .apply(
            normalize_digits
        )
    )

    rows = rows[
        rows[
            "Item Number"
        ].ne("")
    ]

    groups = []

    # Consolidate again here, so even if the user
    # manually creates duplicate rows for the same
    # Costco item, Clover asks for ONE mapping only.
    for (
        item_number,
        group,
    ) in rows.groupby(
        "Item Number",
        sort=False,
    ):

        pack_types = [
            text_value(v)
            for v
            in group[
                "Pack Type"
            ].tolist()
            if text_value(v)
        ]

        unique_packs = sorted(
            set(
                pack_types
            )
        )

        pack_conflict = (
            len(
                unique_packs
            )
            > 1
        )

        pack_type = (
            unique_packs[0]
            if len(
                unique_packs
            )
            == 1
            else ""
        )

        q1 = int(
            pd.to_numeric(
                group[
                    "Qty Purchased"
                ],
                errors="coerce",
            )
            .fillna(0)
            .sum()
        )

        q2 = int(
            pd.to_numeric(
                group[
                    "Qty Returned"
                ],
                errors="coerce",
            )
            .fillna(0)
            .sum()
        )

        purchase = (
            money_round(
                pd.to_numeric(
                    group[
                        "Purchase Total"
                    ],
                    errors="coerce",
                )
                .fillna(0)
                .sum()
            )
            or 0.0
        )

        credit = (
            money_round(
                pd.to_numeric(
                    group[
                        "Credit / Adjustment"
                    ],
                    errors="coerce",
                )
                .fillna(0)
                .sum()
            )
            or 0.0
        )

        pack_size = (
            PACK_TYPES.get(
                pack_type
            )
        )

        proposed_cost = (
            divide_currency(
                purchase,
                q1
                * pack_size,
            )
            if (
                q1 > 0
                and pack_size
            )
            else None
        )

        groups.append(
            {
                "Item Number":
                    item_number,

                "Description":
                    choose_description(
                        group[
                            "Description"
                        ].tolist()
                    ),

                "Qty Purchased":
                    q1,

                "Qty Returned":
                    q2,

                "Purchase Total":
                    purchase,

                "Credit / Adjustment":
                    credit,

                "Net Amount":
                    (
                        money_round(
                            purchase
                            + credit
                        )
                        or 0.0
                    ),

                "Pack Type":
                    pack_type,

                "Pack Conflict":
                    pack_conflict,

                "Cost Variance":
                    any(
                        "YES"
                        in text_value(v)
                        .upper()

                        for v
                        in group[
                            "Cost Variance?"
                        ].tolist()
                    ),

                "Proposed Unit Cost":
                    proposed_cost,
            }
        )

    return pd.DataFrame(
        groups
    )


# ============================================================
# CLOVER SEARCH
# ============================================================

def search_clover(
    df,
    search_text,
    fallback_description="",
    limit=40,
):

    raw = (
        text_value(
            search_text
        )
        or text_value(
            fallback_description
        )
    )

    qname = (
        normalize_name(
            raw
        )
    )

    qdigits = (
        normalize_digits(
            raw
        )
    )

    scored = []

    for i, row in (
        df.iterrows()
    ):

        name = text_value(
            row.get(
                "_Name Clean"
            )
        )

        product = text_value(
            row.get(
                "_Product Code Clean"
            )
        )

        sku = text_value(
            row.get(
                "_SKU Clean"
            )
        )

        score = 0.0

        if (
            qdigits
            and qdigits
            in product
        ):
            score = max(
                score,
                5.0,
            )

        if (
            qname
            and qname
            in name
        ):
            score = max(
                score,
                4.0,
            )

        if (
            qname
            and qname
            in sku
        ):
            score = max(
                score,
                3.5,
            )

        if qname:

            tokens = [
                t
                for t
                in qname.split()
                if t
            ]

            if tokens:

                score = max(
                    score,
                    2.0
                    * sum(
                        1
                        for t
                        in tokens
                        if t in name
                    )
                    / len(
                        tokens
                    ),
                )

            score = max(
                score,
                difflib
                .SequenceMatcher(
                    None,
                    qname,
                    name,
                )
                .ratio(),
            )

        if score > 0:

            scored.append(
                (
                    score,
                    i,
                )
            )

    scored.sort(
        key=lambda x: (
            x[0],
            -x[1],
        ),
        reverse=True,
    )

    return [
        i
        for _, i
        in scored[
            :limit
        ]
    ]


def clover_label(
    row,
):

    name = text_value(
        row.get(
            "Name"
        )
    )

    retail = to_number(
        row.get(
            "Price"
        )
    )

    cost = to_number(
        row.get(
            "Cost"
        )
    )

    parts = [
        name,

        (
            f"Retail ${retail:.2f}"
            if retail
            is not None
            else "Retail —"
        ),

        (
            f"Cost ${cost:.2f}"
            if (
                cost is not None
                and cost > 0
            )
            else "Cost —"
        ),
    ]

    if text_value(
        row.get(
            "Product Code"
        )
    ):

        parts.append(
            f"UPC "
            f"{text_value(row.get('Product Code'))}"
        )

    if text_value(
        row.get(
            "Clover ID"
        )
    ):

        parts.append(
            text_value(
                row.get(
                    "Clover ID"
                )
            )
        )

    return "  |  ".join(
        parts
    )


def match_index_from_key(
    match_key,
    df,
):

    if not match_key:
        return None

    matches = (
        df.index[
            df[
                "_Match Key"
            ].eq(
                match_key
            )
        ]
        .tolist()
    )

    return (
        matches[0]
        if matches
        else None
    )


def active_selected_match_keys(
    item_numbers,
):

    out = {}

    for item in item_numbers:

        action = (
            st.session_state
            .get(
                f"action_{item}",
                "Match existing",
            )
        )

        if (
            action
            == "Match existing"
        ):

            value = (
                st.session_state
                .get(
                    f"match_{item}"
                )
            )

            if value:
                out[item] = value

    return out


# ============================================================
# SAFE CLOVER OUTPUT
# ============================================================

def copy_cell_style(
    source,
    target,
):

    if not source.has_style:
        return

    target.font = copy(
        source.font
    )

    target.fill = copy(
        source.fill
    )

    target.border = copy(
        source.border
    )

    target.alignment = copy(
        source.alignment
    )

    target.protection = copy(
        source.protection
    )

    target.number_format = (
        source.number_format
    )


def build_clover_output(
    clover_bytes,
    clover_df,
    header_index,
    plans,
):

    errors = []

    wb = load_workbook(
        io.BytesIO(
            clover_bytes
        ),
        data_only=False,
    )

    if (
        "Items"
        not in wb.sheetnames
    ):

        return (
            None,
            [
                "The Clover workbook does not "
                "contain an Items sheet."
            ],
            None,
        )

    ws = wb[
        "Items"
    ]

    header_row = (
        header_index
        + 1
    )

    headers = {
        text_value(
            ws.cell(
                header_row,
                c,
            ).value
        ): c

        for c
        in range(
            1,
            ws.max_column
            + 1,
        )

        if text_value(
            ws.cell(
                header_row,
                c,
            ).value
        )
    }

    missing = [
        h
        for h in (
            "Clover ID",
            "Name",
            "Price",
            "Cost",
            "Product Code",
        )
        if h
        not in headers
    ]

    if missing:

        return (
            None,
            [
                "Items sheet is missing: "
                + ", ".join(
                    missing
                )
            ],
            None,
        )

    original_count = (
        len(
            clover_df
        )
    )

    last_row = int(
        clover_df[
            "_Excel Row"
        ].max()
    )

    # Snapshot all existing item cells.
    snapshot = {
        (
            r,
            c,
        ):
            ws.cell(
                r,
                c,
            ).value

        for r
        in clover_df[
            "_Excel Row"
        ].astype(int)

        for c
        in headers.values()
    }

    allowed = set()

    # Existing Clover IDs must not be duplicated.
    ids = [
        text_value(v)
        for v
        in clover_df[
            "Clover ID"
        ].tolist()
        if text_value(v)
    ]

    errors += [
        f"Source Clover ID "
        f"{v} appears more than once."

        for v, count
        in Counter(
            ids
        ).items()

        if count > 1
    ]

    existing_upcs = {
        normalize_digits(v)

        for v
        in clover_df.get(
            "Product Code",
            pd.Series(
                dtype=object
            ),
        ).tolist()

        if normalize_digits(v)
    }

    new_upcs = []

    existing_matches = [
        p[
            "Match Key"
        ]
        for p
        in plans
        if p[
            "Action"
        ]
        == "Match existing"
    ]

    if (
        len(
            existing_matches
        )
        != len(
            set(
                existing_matches
            )
        )
    ):

        errors.append(
            "The same Clover item is mapped "
            "to more than one receipt item."
        )

    # Validate new items.
    for p in plans:

        if (
            p["Action"]
            != "Create new"
        ):
            continue

        upc = (
            normalize_digits(
                p.get(
                    "Product Code"
                )
            )
        )

        name = text_value(
            p.get(
                "Name"
            )
        )

        retail = to_number(
            p.get(
                "Final Retail"
            )
        )

        cost = to_number(
            p.get(
                "Final Cost"
            )
        )

        if not name:

            errors.append(
                f"Costco item "
                f"{p['Item Number']}: "
                "new Clover item name "
                "is missing."
            )

        if len(upc) < 8:

            errors.append(
                f"{name or p['Item Number']}: "
                "a valid UPC/Product Code "
                "is required for a new item."
            )

        if (
            upc
            in existing_upcs
        ):

            errors.append(
                f"{name or p['Item Number']}: "
                f"UPC {upc} already exists "
                "in Clover."
            )

        if upc:
            new_upcs.append(
                upc
            )

        if (
            retail is None
            or retail <= 0
        ):

            errors.append(
                f"{name or p['Item Number']}: "
                "retail price is required."
            )

        if (
            cost is None
            or cost <= 0
        ):

            errors.append(
                f"{name or p['Item Number']}: "
                "valid cost is required."
            )

        if not text_value(
            p.get(
                "Category"
            )
        ):

            errors.append(
                f"{name or p['Item Number']}: "
                "category is required."
            )

        if (
            text_value(
                p.get(
                    "Default tax rates?"
                )
            )
            not in {
                "Yes",
                "No",
            }
        ):

            errors.append(
                f"{name or p['Item Number']}: "
                "tax setting is required."
            )

    errors += [
        f"New UPC "
        f"{v} is used more than once."

        for v, count
        in Counter(
            new_upcs
        ).items()

        if count > 1
    ]

    if errors:

        return (
            None,
            errors,
            None,
        )

    # Existing item updates.
    for p in plans:

        if (
            p["Action"]
            != "Match existing"
        ):
            continue

        row = int(
            p[
                "Excel Row"
            ]
        )

        final_cost = (
            to_number(
                p.get(
                    "Final Cost"
                )
            )
        )

        final_retail = (
            to_number(
                p.get(
                    "Final Retail"
                )
            )
        )

        if (
            final_cost is not None
            and final_cost > 0
            and values_different(
                snapshot[
                    (
                        row,
                        headers[
                            "Cost"
                        ],
                    )
                ],
                final_cost,
            )
        ):

            ws.cell(
                row,
                headers[
                    "Cost"
                ],
            ).value = (
                money_round(
                    final_cost
                )
            )

            allowed.add(
                (
                    row,
                    headers[
                        "Cost"
                    ],
                )
            )

        # Retail only changes if the user
        # explicitly edited it.
        if (
            p.get(
                "Update Retail"
            )
            and final_retail
            is not None
            and values_different(
                snapshot[
                    (
                        row,
                        headers[
                            "Price"
                        ],
                    )
                ],
                final_retail,
            )
        ):

            ws.cell(
                row,
                headers[
                    "Price"
                ],
            ).value = (
                money_round(
                    final_retail
                )
            )

            allowed.add(
                (
                    row,
                    headers[
                        "Price"
                    ],
                )
            )

    # Append brand-new items only.
    new_count = 0
    template_row = last_row

    for p in plans:

        if (
            p["Action"]
            != "Create new"
        ):
            continue

        new_count += 1

        row = (
            last_row
            + new_count
        )

        for c in range(
            1,
            ws.max_column
            + 1,
        ):

            copy_cell_style(
                ws.cell(
                    template_row,
                    c,
                ),
                ws.cell(
                    row,
                    c,
                ),
            )

        ws.row_dimensions[
            row
        ].height = (
            ws.row_dimensions[
                template_row
            ].height
        )

        values = {
            "Clover ID":
                None,

            "Name":
                text_value(
                    p.get(
                        "Name"
                    )
                ),

            "Alternate Name":
                None,

            "Description":
                None,

            "Price":
                money_round(
                    p.get(
                        "Final Retail"
                    )
                ),

            "Price Type":
                "Fixed",

            "Price Unit":
                None,

            "Cost":
                money_round(
                    p.get(
                        "Final Cost"
                    )
                ),

            "Product Code":
                normalize_digits(
                    p.get(
                        "Product Code"
                    )
                ),

            "SKU":
                "",

            "Quantity":
                None,

            "Hidden?":
                "No",

            "Default tax rates?":
                text_value(
                    p.get(
                        "Default tax rates?"
                    )
                ),

            "Non-revenue item?":
                "No",

            "Printer Labels":
                (
                    text_value(
                        p.get(
                            "Printer Labels"
                        )
                    )
                    or None
                ),

            "Modifier Groups":
                None,

            "Categories":
                text_value(
                    p.get(
                        "Category"
                    )
                ),

            # New item only.
            # Existing Tax Rates are NEVER cleared.
            "Tax Rates":
                None,

            "Variant Attribute":
                None,

            "Variant Option":
                None,
        }

        for h, v in (
            values.items()
        ):

            if h in headers:

                ws.cell(
                    row,
                    headers[h],
                ).value = v

    # Hard safety check:
    # every existing cell must remain identical
    # except explicitly authorized Cost/Price cells.
    for (
        (
            r,
            c,
        ),
        original,
    ) in snapshot.items():

        if (
            r,
            c,
        ) in allowed:
            continue

        if (
            ws.cell(
                r,
                c,
            ).value
            != original
        ):

            header_name = next(
                (
                    name
                    for name, col
                    in headers.items()
                    if col == c
                ),
                f"Column {c}",
            )

            return (
                None,
                [
                    f"Safety check failed: "
                    f"existing row {r}, "
                    f"column '{header_name}' "
                    "changed unexpectedly."
                ],
                None,
            )

    output = io.BytesIO()

    wb.save(
        output
    )

    data = output.getvalue()

    if (
        len(data)
        > 10
        * 1024
        * 1024
    ):

        return (
            None,
            [
                "Generated file exceeds "
                "Clover's 10 MB upload limit."
            ],
            None,
        )

    diagnostics = {
        "Original items":
            original_count,

        "Existing updated":
            len(
                {
                    r
                    for r, _
                    in allowed
                }
            ),

        "New items":
            new_count,

        "Skipped receipt items":
            sum(
                1
                for p
                in plans
                if p[
                    "Action"
                ]
                == "Skip"
            ),

        "Final items":
            original_count
            + new_count,
    }

    return (
        data,
        [],
        diagnostics,
    )


# ============================================================
# PROCESS CURRENT RECEIPT
# ============================================================

def process_current_receipt(
    receipt_files,
    clover_file,
    mode,
    rotation,
):

    if clover_file is None:

        raise ValueError(
            "Upload the current Clover "
            "inventory export first."
        )

    if not receipt_files:

        raise ValueError(
            "Upload at least one receipt "
            "photo or PDF."
        )

    clover_bytes = (
        clover_file
        .getvalue()
    )

    (
        clover_df,
        header_index,
    ) = load_clover(
        clover_bytes
    )

    raw_ocr = (
        read_receipt_files(
            receipt_files,
            mode,
            rotation,
        )
    )

    rows = (
        parse_costco_receipt(
            raw_ocr
        )
    )

    st.session_state.clover_bytes = (
        clover_bytes
    )

    st.session_state.clover_df = (
        clover_df
    )

    st.session_state.clover_header_index = (
        header_index
    )

    st.session_state.raw_ocr = (
        raw_ocr
    )

    st.session_state.receipt_rows = (
        rows
    )

    st.session_state.ocr_mode = (
        mode
    )

    st.session_state.processed = (
        True
    )

    clear_mapping_state()
    clear_prepared_output()

    st.session_state.pop(
        "receipt_editor",
        None,
    )


# ============================================================
# HEADER
# ============================================================

header_left, header_right = (
    st.columns(
        [
            5,
            1,
        ]
    )
)

with header_right:

    theme = st.selectbox(
        "Theme",
        [
            "Dark",
            "Light",
        ],
        key="theme_mode",
    )


apply_theme(
    theme
)


with header_left:

    st.title(
        "Vendor Invoice Processor"
    )

    st.caption(
        "Costco receipt OCR → "
        "total reconciliation → "
        "cost / margin review → "
        "safe Clover file"
    )


# ============================================================
# INPUT
# ============================================================

with st.container(
    border=True
):

    st.subheader(
        "1. Current files"
    )

    left, right = (
        st.columns(2)
    )

    with left:

        clover_file = (
            st.file_uploader(
                "Current Clover inventory export",
                type=[
                    "xlsx"
                ],
                key="clover_upload",
                on_change=(
                    reset_processed_data
                ),
            )
        )

    with right:

        receipt_files = (
            st.file_uploader(
                "Costco receipt pages / "
                "photos / scans",
                type=[
                    "jpg",
                    "jpeg",
                    "png",
                    "pdf",
                    "tif",
                    "tiff",
                ],
                accept_multiple_files=True,
                key="receipt_upload",
                on_change=(
                    reset_processed_data
                ),
                help=(
                    "Upload all pages/photos "
                    "belonging to the same "
                    "Costco receipt."
                ),
            )
        )


with st.container(
    border=True
):

    st.subheader(
        "2. Safety controls"
    )

    c1, c2, c3 = (
        st.columns(3)
    )

    with c1:

        expected_total = (
            st.number_input(
                "Final receipt total "
                "(required)",
                value=0.0,
                step=0.01,
                format="%.2f",
                key=(
                    "expected_total_input"
                ),
            )
        )

    with c2:

        expected_count_text = (
            st.text_input(
                "Total items/cartons "
                "(optional)",
                placeholder=(
                    "Example: 6"
                ),
                key=(
                    "expected_count_input"
                ),
            )
        )

    with c3:

        rotation = (
            st.selectbox(
                "Rotate receipt before OCR",
                [
                    0,
                    90,
                    180,
                    270,
                ],
                format_func=(
                    lambda v:
                    "No rotation"
                    if v == 0
                    else f"{v}°"
                ),
                key=(
                    "rotation_input"
                ),
            )
        )


ready = (
    clover_file
    is not None
    and bool(
        receipt_files
    )
    and expected_total
    != 0
)


if st.button(
    "Read & Reconcile Receipt",
    type="primary",
    use_container_width=True,
    disabled=not ready,
):

    try:

        with st.status(
            "Reading Costco receipt...",
            expanded=True,
        ) as status:

            status.write(
                "Reading current Clover "
                "inventory..."
            )

            status.write(
                "Running standard OCR..."
            )

            process_current_receipt(
                receipt_files,
                clover_file,
                "Standard",
                rotation,
            )

            status.write(
                "Consolidating repeated "
                "Costco item numbers..."
            )

            status.update(
                label=(
                    "Receipt ready for review"
                ),
                state="complete",
                expanded=False,
            )

        st.rerun()

    except Exception as exc:

        st.error(
            f"Receipt processing stopped: "
            f"{exc}"
        )


# ============================================================
# POST-OCR WORKFLOW
# ============================================================

if st.session_state.processed:

    st.write("")

    r1, r2, r3 = (
        st.columns(3)
    )

    with r1:

        if st.button(
            "Re-read Receipt",
            use_container_width=True,
        ):

            try:

                with st.status(
                    "Re-reading receipt...",
                    expanded=True,
                ) as status:

                    process_current_receipt(
                        receipt_files,
                        clover_file,
                        "Standard",
                        rotation,
                    )

                    status.update(
                        label=(
                            "Standard OCR complete"
                        ),
                        state="complete",
                        expanded=False,
                    )

                st.rerun()

            except Exception as exc:

                st.error(
                    f"Re-read failed: "
                    f"{exc}"
                )

    with r2:

        if st.button(
            "Re-read with Strong OCR",
            use_container_width=True,
        ):

            try:

                with st.status(
                    "Running stronger OCR...",
                    expanded=True,
                ) as status:

                    status.write(
                        "Enhancing contrast and "
                        "reading long pages in "
                        "smaller sections..."
                    )

                    process_current_receipt(
                        receipt_files,
                        clover_file,
                        "Strong",
                        rotation,
                    )

                    status.update(
                        label=(
                            "Strong OCR complete"
                        ),
                        state="complete",
                        expanded=False,
                    )

                st.rerun()

            except Exception as exc:

                st.error(
                    f"Strong re-read failed: "
                    f"{exc}"
                )

    with r3:

        if st.button(
            "Clear Review & Start Over",
            use_container_width=True,
        ):

            reset_processed_data()
            st.rerun()


    st.caption(
        f"Current OCR mode: "
        f"{st.session_state.ocr_mode}"
    )


    (
        tab_receipt,
        tab_mapping,
        tab_output,
        tab_ocr,
    ) = st.tabs(
        [
            "Receipt & Reconciliation",
            "Clover Mapping",
            "Final Clover File",
            "OCR Details",
        ]
    )


    # ========================================================
    # RECEIPT REVIEW
    # ========================================================

    with tab_receipt:

        st.subheader(
            "3. Review receipt lines"
        )

        st.caption(
            "Repeated Costco item numbers "
            "are consolidated. Add an "
            "Adjustment row for a discount, "
            "fee, or other amount OCR did "
            "not identify as an item."
        )

        source_rows = (
            st.session_state.receipt_rows
            if (
                st.session_state.receipt_rows
                is not None
            )
            else pd.DataFrame(
                columns=(
                    RECEIPT_COLUMNS
                )
            )
        )

        edited_rows = (
            st.data_editor(
                source_rows,
                use_container_width=True,
                hide_index=True,
                num_rows="dynamic",
                key="receipt_editor",
                column_config={

                    "Row Type":
                        st.column_config
                        .SelectboxColumn(
                            "Row Type",
                            options=[
                                "Item",
                                "Adjustment",
                            ],
                            required=True,
                        ),

                    "Item Number":
                        st.column_config
                        .TextColumn(
                            "Costco Item #",
                            help=(
                                "Required for Item "
                                "rows; blank is allowed "
                                "for Adjustment rows."
                            ),
                        ),

                    "Description":
                        st.column_config
                        .TextColumn(
                            "Description",
                            width="large",
                        ),

                    "Qty Purchased":
                        st.column_config
                        .NumberColumn(
                            "Qty Bought",
                            min_value=0,
                            step=1,
                        ),

                    "Qty Returned":
                        st.column_config
                        .NumberColumn(
                            "Qty Returned",
                            min_value=0,
                            step=1,
                        ),

                    "Purchase Total":
                        st.column_config
                        .NumberColumn(
                            "Purchase Total",
                            format="$%.2f",
                        ),

                    "Credit / Adjustment":
                        st.column_config
                        .NumberColumn(
                            "Credit / Adjustment",
                            format="$%.2f",
                            help=(
                                "Negative for return/"
                                "discount/credit; "
                                "positive for fee."
                            ),
                        ),

                    "Net Amount":
                        st.column_config
                        .NumberColumn(
                            "Net Amount",
                            format="$%.2f",
                            disabled=True,
                        ),

                    "Pack Type":
                        st.column_config
                        .SelectboxColumn(
                            "Pack Type",
                            options=list(
                                PACK_TYPES.keys()
                            ),
                        ),

                    "Cost Variance?":
                        st.column_config
                        .TextColumn(
                            "Cost Variance?",
                            disabled=True,
                        ),
                },
            )
        )

        edited_rows = (
            recompute_receipt_rows(
                edited_rows
            )
        )

        st.session_state.receipt_rows = (
            edited_rows
        )

        calculated_total = (
            calculated_receipt_total(
                edited_rows
            )
        )

        calculated_count = (
            calculated_item_count(
                edited_rows
            )
        )

        difference = (
            money_round(
                calculated_total
                - expected_total
            )
            or 0.0
        )

        m1, m2, m3, m4 = (
            st.columns(4)
        )

        m1.metric(
            "Your receipt total",
            f"${expected_total:,.2f}",
        )

        m2.metric(
            "App calculated total",
            f"${calculated_total:,.2f}",
        )

        m3.metric(
            "Difference",
            f"${difference:,.2f}",
        )

        m4.metric(
            "Calculated item count",
            calculated_count,
        )

        totals_match = (
            abs(
                difference
            )
            <= 0.01
        )

        entered_count = None
        count_error = False

        if (
            expected_count_text
            .strip()
        ):

            try:

                entered_count = int(
                    expected_count_text
                    .strip()
                )

            except ValueError:
                count_error = True

        count_match = (
            not count_error
            and (
                entered_count
                is None
                or entered_count
                == calculated_count
            )
        )

        if totals_match:

            st.success(
                "Receipt total matches "
                "the user-entered total."
            )

        else:

            st.error(
                "Receipt total does not match. "
                "Fix the extracted lines or "
                "add an Adjustment row before "
                "continuing."
            )

        if count_error:

            st.error(
                "Optional item count must "
                "be a whole number."
            )

        elif (
            entered_count
            is not None
        ):

            if count_match:

                st.success(
                    f"Optional item count "
                    f"also matches: "
                    f"{calculated_count}."
                )

            else:

                st.error(
                    f"Item count does not match. "
                    f"You entered "
                    f"{entered_count}; "
                    f"the app calculated "
                    f"{calculated_count}."
                )

        receipt_row_errors = (
            validate_receipt_rows(
                edited_rows
            )
        )

        if receipt_row_errors:

            with st.expander(
                f"Receipt rows needing "
                f"attention "
                f"({len(receipt_row_errors)})",
                expanded=True,
            ):

                for error in (
                    receipt_row_errors
                ):

                    st.warning(
                        error
                    )

        product_groups = (
            build_product_groups(
                edited_rows
            )
        )

        if product_groups.empty:

            st.warning(
                "No valid Costco item rows "
                "are available yet. You can "
                "add/correct rows above manually."
            )


    # ========================================================
    # CLOVER MAPPING
    # ========================================================

    mapping_plans = []
    mapping_errors = []


    with tab_mapping:

        st.subheader(
            "4. Match receipt items to Clover"
        )

        if (
            not totals_match
            or not count_match
            or receipt_row_errors
        ):

            st.warning(
                "Clover mapping is blocked "
                "until the receipt reconciliation "
                "and receipt-row checks pass."
            )

        elif product_groups.empty:

            st.warning(
                "There are no Costco item rows "
                "to map."
            )

        else:

            st.caption(
                "Search by Clover item name, "
                "UPC/Product Code, or SKU. "
                "Once selected, a Clover item "
                "is removed from the other "
                "match lists."
            )

            clover_df = (
                st.session_state
                .clover_df
            )

            categories = (
                clover_categories(
                    clover_df
                )
            )

            item_numbers = (
                product_groups[
                    "Item Number"
                ]
                .astype(str)
                .tolist()
            )

            active_matches = (
                active_selected_match_keys(
                    item_numbers
                )
            )

            for _, item in (
                product_groups
                .iterrows()
            ):

                item_number = (
                    text_value(
                        item[
                            "Item Number"
                        ]
                    )
                )

                description = (
                    text_value(
                        item[
                            "Description"
                        ]
                    )
                )

                q1 = int(
                    to_number(
                        item[
                            "Qty Purchased"
                        ]
                    )
                    or 0
                )

                q2 = int(
                    to_number(
                        item[
                            "Qty Returned"
                        ]
                    )
                    or 0
                )

                purchase = (
                    to_number(
                        item[
                            "Purchase Total"
                        ]
                    )
                    or 0.0
                )

                credit = (
                    to_number(
                        item[
                            "Credit / Adjustment"
                        ]
                    )
                    or 0.0
                )

                proposed_cost = (
                    to_number(
                        item[
                            "Proposed Unit Cost"
                        ]
                    )
                )

                pack_conflict = bool(
                    item[
                        "Pack Conflict"
                    ]
                )

                cost_variance = bool(
                    item[
                        "Cost Variance"
                    ]
                )

                action_key = (
                    f"action_"
                    f"{item_number}"
                )

                search_key = (
                    f"search_"
                    f"{item_number}"
                )

                match_widget = (
                    f"match_"
                    f"{item_number}"
                )


                with st.container(
                    border=True
                ):

                    st.markdown(
                        f"### {description}"
                    )

                    a, b, c, d, e = (
                        st.columns(5)
                    )

                    a.metric(
                        "Costco Item #",
                        item_number,
                    )

                    b.metric(
                        "Qty Bought",
                        q1,
                    )

                    c.metric(
                        "Qty Returned",
                        q2,
                    )

                    d.metric(
                        "Purchase Total",
                        f"${purchase:,.2f}",
                    )

                    e.metric(
                        "Calculated Clover Cost",
                        (
                            f"${proposed_cost:,.2f}"
                            if proposed_cost
                            is not None
                            else "—"
                        ),
                    )

                    if credit:

                        st.info(
                            f"Returns / credits on "
                            f"this receipt: "
                            f"${credit:,.2f}. "
                            "They are not used to "
                            "calculate Clover cost."
                        )

                    if pack_conflict:

                        st.error(
                            "This item has conflicting "
                            "Pack Type values. Return "
                            "to Receipt & Reconciliation "
                            "and make them consistent."
                        )

                        mapping_errors.append(
                            f"Costco item "
                            f"{item_number}: "
                            "conflicting Pack Type values."
                        )

                    if cost_variance:

                        st.warning(
                            "This item appeared at more "
                            "than one positive package "
                            "price. The proposed Clover "
                            "cost uses weighted average "
                            "purchase cost."
                        )

                        confirmed = (
                            st.checkbox(
                                "I reviewed the different "
                                "package prices and want "
                                "to use the average "
                                "purchase cost",
                                key=(
                                    f"variance_confirm_"
                                    f"{item_number}"
                                ),
                            )
                        )

                        if not confirmed:

                            mapping_errors.append(
                                f"Costco item "
                                f"{item_number}: "
                                "variable package cost "
                                "has not been confirmed."
                            )

                    if q1 <= 0:

                        st.warning(
                            "No positive purchase quantity "
                            "exists. A return/credit alone "
                            "will never overwrite Clover cost."
                        )


                    action = (
                        st.selectbox(
                            "What should Clover do "
                            "with this receipt item?",
                            [
                                "Match existing",
                                "Create new",
                                "Skip",
                            ],
                            key=action_key,
                        )
                    )


                    # ----------------------------------------
                    # SKIP
                    # ----------------------------------------

                    if action == "Skip":

                        st.warning(
                            "This receipt item will "
                            "reconcile normally but "
                            "will not change Clover."
                        )

                        mapping_plans.append(
                            {
                                "Action":
                                    "Skip",

                                "Item Number":
                                    item_number,

                                "Description":
                                    description,
                            }
                        )

                        continue


                    # ----------------------------------------
                    # MATCH EXISTING
                    # ----------------------------------------

                    if (
                        action
                        == "Match existing"
                    ):

                        search_text = (
                            st.text_input(
                                "Search Clover by "
                                "name, UPC or SKU",
                                placeholder=(
                                    description
                                ),
                                key=search_key,
                            )
                        )

                        candidates = (
                            search_clover(
                                clover_df,
                                search_text,
                                description,
                            )
                        )

                        current_match = (
                            st.session_state
                            .get(
                                match_widget
                            )
                        )

                        current_index = (
                            match_index_from_key(
                                current_match,
                                clover_df,
                            )
                        )

                        if (
                            current_index
                            is not None
                            and current_index
                            not in candidates
                        ):

                            candidates.insert(
                                0,
                                current_index,
                            )

                        used_elsewhere = {
                            match_key

                            for (
                                other_item,
                                match_key,
                            )
                            in active_matches.items()

                            if (
                                other_item
                                != item_number
                                and match_key
                            )
                        }

                        available = [
                            i

                            for i
                            in candidates

                            if (
                                clover_df.loc[
                                    i,
                                    "_Match Key",
                                ]
                                not in used_elsewhere

                                or clover_df.loc[
                                    i,
                                    "_Match Key",
                                ]
                                == current_match
                            )
                        ]

                        options = [
                            None
                        ] + [
                            clover_df.loc[
                                i,
                                "_Match Key",
                            ]

                            for i
                            in available
                        ]

                        if (
                            current_match
                            and current_match
                            not in options
                        ):

                            options.insert(
                                1,
                                current_match,
                            )


                        def format_match(
                            key,
                        ):

                            if key is None:
                                return (
                                    "— Select Clover item —"
                                )

                            idx = (
                                match_index_from_key(
                                    key,
                                    clover_df,
                                )
                            )

                            if idx is None:
                                return "Unavailable"

                            return clover_label(
                                clover_df.loc[
                                    idx
                                ]
                            )


                        selected = (
                            st.selectbox(
                                "Clover match",
                                options,
                                key=match_widget,
                                format_func=(
                                    format_match
                                ),
                            )
                        )

                        if selected is None:

                            mapping_errors.append(
                                f"Costco item "
                                f"{item_number}: "
                                "no Clover item selected."
                            )

                            st.info(
                                "Select the correct "
                                "Clover item or choose "
                                "Create new / Skip."
                            )

                            continue


                        selected_index = (
                            match_index_from_key(
                                selected,
                                clover_df,
                            )
                        )

                        if (
                            selected_index
                            is None
                        ):

                            mapping_errors.append(
                                f"Costco item "
                                f"{item_number}: "
                                "selected Clover match "
                                "is unavailable."
                            )

                            continue


                        clover_row = (
                            clover_df.loc[
                                selected_index
                            ]
                        )

                        current_cost = (
                            to_number(
                                clover_row.get(
                                    "Cost"
                                )
                            )
                        )

                        current_retail = (
                            to_number(
                                clover_row.get(
                                    "Price"
                                )
                            )
                        )

                        if (
                            current_cost
                            is not None
                            and current_cost
                            <= 0
                        ):

                            current_cost = None


                        old_margin = (
                            margin_percent(
                                current_cost,
                                current_retail,
                            )
                        )

                        new_margin = (
                            margin_percent(
                                proposed_cost,
                                current_retail,
                            )
                        )

                        if (
                            not pd.isna(
                                old_margin
                            )
                            and not pd.isna(
                                new_margin
                            )
                        ):

                            delta = round(
                                new_margin
                                - old_margin,
                                1,
                            )

                        else:
                            delta = np.nan


                        m1, m2, m3, m4 = (
                            st.columns(4)
                        )

                        m1.metric(
                            "Current Clover Cost",
                            (
                                f"${current_cost:,.2f}"
                                if current_cost
                                is not None
                                else "—"
                            ),
                        )

                        m2.metric(
                            "Current Retail",
                            (
                                f"${current_retail:,.2f}"
                                if current_retail
                                is not None
                                else "—"
                            ),
                        )

                        m3.metric(
                            "Old Margin",
                            (
                                f"{old_margin:.1f}%"
                                if not pd.isna(
                                    old_margin
                                )
                                else "—"
                            ),
                        )

                        m4.metric(
                            "Margin with New Cost",
                            (
                                f"{new_margin:.1f}%"
                                if not pd.isna(
                                    new_margin
                                )
                                else "—"
                            ),
                            delta=(
                                f"{delta:+.1f} pts"
                                if not pd.isna(
                                    delta
                                )
                                else None
                            ),
                        )


                        default_cost_action = (
                            "Use invoice cost"
                            if (
                                proposed_cost
                                is not None
                                and proposed_cost
                                > 0
                            )
                            else "Keep Clover cost"
                        )


                        cost_state_key = (
                            f"cost_action_"
                            f"{item_number}"
                        )

                        if (
                            cost_state_key
                            not in st.session_state
                        ):

                            st.session_state[
                                cost_state_key
                            ] = (
                                default_cost_action
                            )


                        cost_action = (
                            st.selectbox(
                                "Cost action",
                                [
                                    "Use invoice cost",
                                    "Keep Clover cost",
                                ],
                                key=(
                                    cost_state_key
                                ),
                                disabled=(
                                    proposed_cost
                                    is None
                                    or proposed_cost
                                    <= 0
                                ),
                            )
                        )


                        final_cost = (
                            proposed_cost
                            if (
                                cost_action
                                == "Use invoice cost"
                                and proposed_cost
                                is not None
                                and proposed_cost
                                > 0
                            )
                            else current_cost
                        )


                        retail_state_key = (
                            f"retail_"
                            f"{item_number}"
                        )

                        if (
                            retail_state_key
                            not in st.session_state
                        ):

                            st.session_state[
                                retail_state_key
                            ] = float(
                                current_retail
                                or 0.0
                            )


                        final_retail = (
                            st.number_input(
                                "Final retail price",
                                min_value=0.0,
                                step=0.01,
                                format="%.2f",
                                key=(
                                    retail_state_key
                                ),
                                help=(
                                    "Defaults to current "
                                    "Clover retail and "
                                    "changes only if you "
                                    "edit it."
                                ),
                            )
                        )


                        if (
                            proposed_cost
                            is not None
                            and current_cost
                            is not None
                        ):

                            if (
                                proposed_cost
                                > current_cost
                                + 0.004
                            ):

                                st.warning(
                                    "Cost increased. "
                                    "New cost is proposed "
                                    "automatically; retail "
                                    "remains unchanged unless "
                                    "you edit it."
                                )

                                preserve = (
                                    retail_for_margin(
                                        proposed_cost,
                                        old_margin,
                                    )
                                )

                                if (
                                    preserve
                                    is not None
                                ):

                                    st.info(
                                        f"Retail to approximately "
                                        f"preserve the old "
                                        f"{old_margin:.1f}% margin: "
                                        f"${preserve:.2f}. "
                                        "Reference only; retail "
                                        "is not automatic."
                                    )

                            elif (
                                proposed_cost
                                < current_cost
                                - 0.004
                            ):

                                st.info(
                                    "Cost decreased. Lower "
                                    "positive invoice cost is "
                                    "proposed and flagged as a "
                                    "possible promotion. Retail "
                                    "is not lowered automatically."
                                )


                        elif (
                            proposed_cost
                            is not None
                            and current_cost
                            is None
                        ):

                            st.warning(
                                "Clover cost is missing. "
                                "Old margin is unknown; "
                                "positive invoice cost "
                                "is proposed."
                            )


                        if (
                            current_retail
                            is not None
                            and final_retail
                            < current_retail
                            - 0.004
                        ):

                            st.warning(
                                "You are manually lowering "
                                "retail. The app never does "
                                "this automatically."
                            )


                        elif (
                            current_retail
                            is not None
                            and final_retail
                            > current_retail
                            + 0.004
                        ):

                            final_margin = (
                                margin_percent(
                                    final_cost,
                                    final_retail,
                                )
                            )

                            if not pd.isna(
                                final_margin
                            ):

                                st.success(
                                    f"Manual retail increase "
                                    f"selected. Margin at "
                                    f"${final_retail:.2f}: "
                                    f"{final_margin:.1f}%"
                                )

                            else:

                                st.success(
                                    "Manual retail increase "
                                    "selected."
                                )


                        update_retail = (
                            (
                                current_retail
                                is None
                                and final_retail
                                > 0
                            )
                            or (
                                current_retail
                                is not None
                                and values_different(
                                    current_retail,
                                    final_retail,
                                )
                            )
                        )


                        mapping_plans.append(
                            {
                                "Action":
                                    "Match existing",

                                "Item Number":
                                    item_number,

                                "Description":
                                    description,

                                "Match Key":
                                    selected,

                                "Excel Row":
                                    int(
                                        clover_row[
                                            "_Excel Row"
                                        ]
                                    ),

                                "Clover ID":
                                    text_value(
                                        clover_row.get(
                                            "Clover ID"
                                        )
                                    ),

                                "Final Cost":
                                    final_cost,

                                "Final Retail":
                                    final_retail,

                                "Update Retail":
                                    update_retail,
                            }
                        )


                    # ----------------------------------------
                    # CREATE NEW
                    # ----------------------------------------

                    elif (
                        action
                        == "Create new"
                    ):

                        if (
                            proposed_cost
                            is None
                            or proposed_cost
                            <= 0
                        ):

                            mapping_errors.append(
                                f"Costco item "
                                f"{item_number}: "
                                "cannot create a new "
                                "Clover item without "
                                "positive purchase cost."
                            )

                            st.error(
                                "A new Clover item "
                                "requires a positive "
                                "purchase cost."
                            )

                            continue


                        default_category = (
                            "Tobacco"
                            if (
                                "Tobacco"
                                in categories
                            )
                            else categories[0]
                        )


                        new_name = (
                            st.text_input(
                                "New Clover item name",
                                value=description,
                                key=(
                                    f"new_name_"
                                    f"{item_number}"
                                ),
                            )
                        )


                        new_upc = (
                            st.text_input(
                                "UPC / Product Code",
                                key=(
                                    f"new_upc_"
                                    f"{item_number}"
                                ),
                                help=(
                                    "Costco item number is "
                                    "not treated as the UPC. "
                                    "Enter the actual barcode/"
                                    "Product Code."
                                ),
                            )
                        )


                        new_retail_key = (
                            f"new_retail_"
                            f"{item_number}"
                        )

                        if (
                            new_retail_key
                            not in st.session_state
                        ):

                            st.session_state[
                                new_retail_key
                            ] = 0.0


                        new_retail = (
                            st.number_input(
                                "Retail price",
                                min_value=0.0,
                                step=0.01,
                                format="%.2f",
                                key=(
                                    new_retail_key
                                ),
                            )
                        )


                        category = (
                            st.selectbox(
                                "Category",
                                categories,
                                index=(
                                    categories.index(
                                        default_category
                                    )
                                ),
                                key=(
                                    f"new_category_"
                                    f"{item_number}"
                                ),
                            )
                        )


                        taxable = (
                            st.selectbox(
                                "Default tax rates?",
                                [
                                    "Yes",
                                    "No",
                                ],
                                key=(
                                    f"new_tax_"
                                    f"{item_number}"
                                ),
                            )
                        )


                        printer = (
                            st.text_input(
                                "Printer Labels "
                                "(optional)",
                                key=(
                                    f"new_printer_"
                                    f"{item_number}"
                                ),
                            )
                        )


                        new_margin = (
                            margin_percent(
                                proposed_cost,
                                new_retail,
                            )
                        )


                        if (
                            new_retail > 0
                            and not pd.isna(
                                new_margin
                            )
                        ):

                            st.metric(
                                "New item margin",
                                f"{new_margin:.1f}%",
                            )


                        if not text_value(
                            new_name
                        ):

                            mapping_errors.append(
                                f"Costco item "
                                f"{item_number}: "
                                "new item name is missing."
                            )


                        if (
                            len(
                                normalize_digits(
                                    new_upc
                                )
                            )
                            < 8
                        ):

                            mapping_errors.append(
                                f"Costco item "
                                f"{item_number}: "
                                "valid UPC/Product Code "
                                "is required."
                            )


                        if (
                            new_retail
                            <= 0
                        ):

                            mapping_errors.append(
                                f"Costco item "
                                f"{item_number}: "
                                "retail price is required "
                                "for the new item."
                            )


                        mapping_plans.append(
                            {
                                "Action":
                                    "Create new",

                                "Item Number":
                                    item_number,

                                "Description":
                                    description,

                                "Name":
                                    new_name,

                                "Product Code":
                                    new_upc,

                                "Final Cost":
                                    proposed_cost,

                                "Final Retail":
                                    new_retail,

                                "Category":
                                    category,

                                "Default tax rates?":
                                    taxable,

                                "Printer Labels":
                                    printer,
                            }
                        )


            # Final duplicate-mapping protection.
            selected_existing = [
                p[
                    "Match Key"
                ]
                for p
                in mapping_plans
                if (
                    p[
                        "Action"
                    ]
                    == "Match existing"
                    and p.get(
                        "Match Key"
                    )
                )
            ]


            if (
                len(
                    selected_existing
                )
                != len(
                    set(
                        selected_existing
                    )
                )
            ):

                mapping_errors.append(
                    "The same Clover item is mapped "
                    "to more than one Costco item."
                )

                st.error(
                    "Duplicate Clover mapping "
                    "detected. One Clover item "
                    "cannot be used twice."
                )


            mapped = sum(
                1
                for p
                in mapping_plans
                if p[
                    "Action"
                ]
                in {
                    "Match existing",
                    "Create new",
                }
            )

            skipped = sum(
                1
                for p
                in mapping_plans
                if p[
                    "Action"
                ]
                == "Skip"
            )


            s1, s2, s3 = (
                st.columns(3)
            )

            s1.metric(
                "Receipt products",
                len(
                    product_groups
                ),
            )

            s2.metric(
                "Mapped / New",
                mapped,
            )

            s3.metric(
                "Skipped",
                skipped,
            )


            if (
                not mapping_errors
                and len(
                    mapping_plans
                )
                == len(
                    product_groups
                )
            ):

                st.success(
                    "All receipt products have "
                    "a resolved Clover action and "
                    "no Clover item is mapped twice."
                )


            elif mapping_errors:

                with st.expander(
                    f"Items still needing "
                    f"attention "
                    f"({len(mapping_errors)})",
                    expanded=True,
                ):

                    for error in (
                        mapping_errors
                    ):

                        st.warning(
                            error
                        )


    # ========================================================
    # FINAL OUTPUT
    # ========================================================

    with tab_output:

        st.subheader(
            "5. Validate and create "
            "the Clover file"
        )

        st.caption(
            "Output starts from the fresh "
            "Clover export. Existing rows "
            "stay in place; only approved "
            "Cost/Price cells can change. "
            "New items append at the bottom."
        )

        output_errors = []

        if not totals_match:

            output_errors.append(
                "Receipt total does not reconcile."
            )

        if not count_match:

            output_errors.append(
                "Optional item count does "
                "not reconcile."
            )

        output_errors += (
            receipt_row_errors
            + mapping_errors
        )

        if (
            not product_groups.empty
            and len(
                mapping_plans
            )
            != len(
                product_groups
            )
        ):

            output_errors.append(
                "Not every receipt product "
                "has a resolved action."
            )


        if output_errors:

            st.error(
                "Final Clover file is blocked "
                "until these checks are resolved."
            )

            for error in (
                dict.fromkeys(
                    output_errors
                )
            ):

                st.write(
                    f"• {error}"
                )


        # Fingerprint prevents a stale download
        # if the user changes any review decision.
        payload = {
            "expected_total":
                expected_total,

            "expected_count":
                expected_count_text,

            "receipt":
                edited_rows
                .to_dict(
                    "records"
                ),

            "plans":
                mapping_plans,
        }


        fingerprint = (
            hashlib
            .sha256(
                json.dumps(
                    payload,
                    sort_keys=True,
                    default=str,
                ).encode()
            )
            .hexdigest()
        )


        if (
            st.session_state
            .prepared_fingerprint
            != fingerprint
        ):

            clear_prepared_output()


        if st.button(
            "Validate & Prepare Clover File",
            type="primary",
            use_container_width=True,
            disabled=bool(
                output_errors
            ),
        ):

            try:

                with st.status(
                    "Building protected "
                    "Clover file...",
                    expanded=True,
                ) as status:

                    status.write(
                        "Validating mappings "
                        "and new UPCs..."
                    )

                    status.write(
                        "Applying approved "
                        "cost/retail changes only..."
                    )

                    status.write(
                        "Checking unrelated Clover "
                        "cells remain identical..."
                    )

                    (
                        data,
                        errors,
                        diagnostics,
                    ) = (
                        build_clover_output(
                            st.session_state
                            .clover_bytes,

                            st.session_state
                            .clover_df,

                            st.session_state
                            .clover_header_index,

                            mapping_plans,
                        )
                    )


                    if errors:

                        for error in errors:

                            st.error(
                                error
                            )

                        status.update(
                            label=(
                                "Validation failed"
                            ),
                            state="error",
                            expanded=True,
                        )


                    else:

                        st.session_state.prepared_output = (
                            data
                        )

                        st.session_state.prepared_diagnostics = (
                            diagnostics
                        )

                        st.session_state.prepared_filename = (
                            "clover_vendor_IMPORT_"
                            + datetime.now()
                            .strftime(
                                "%Y-%m-%d"
                            )
                            + ".xlsx"
                        )

                        st.session_state.prepared_fingerprint = (
                            fingerprint
                        )

                        status.update(
                            label=(
                                "Clover file ready"
                            ),
                            state="complete",
                            expanded=False,
                        )

                        st.rerun()


            except Exception as exc:

                st.error(
                    f"Clover file generation "
                    f"stopped: {exc}"
                )


        if (
            st.session_state
            .prepared_output
            is not None
            and st.session_state
            .prepared_fingerprint
            == fingerprint
        ):

            d = (
                st.session_state
                .prepared_diagnostics
                or {}
            )

            (
                d1,
                d2,
                d3,
                d4,
                d5,
            ) = st.columns(5)

            d1.metric(
                "Original items",
                d.get(
                    "Original items",
                    0,
                ),
            )

            d2.metric(
                "Existing updated",
                d.get(
                    "Existing updated",
                    0,
                ),
            )

            d3.metric(
                "New items",
                d.get(
                    "New items",
                    0,
                ),
            )

            d4.metric(
                "Skipped",
                d.get(
                    "Skipped receipt items",
                    0,
                ),
            )

            d5.metric(
                "Final items",
                d.get(
                    "Final items",
                    0,
                ),
            )

            st.success(
                "Safety checks passed. "
                "Existing Clover IDs, tax fields, "
                "categories, UPCs, names and "
                "unrelated rows stayed unchanged."
            )

            st.download_button(
                "Download Verified Clover File",
                data=(
                    st.session_state
                    .prepared_output
                ),
                file_name=(
                    st.session_state
                    .prepared_filename
                ),
                mime=(
                    "application/"
                    "vnd.openxmlformats-officedocument."
                    "spreadsheetml.sheet"
                ),
                use_container_width=True,
            )


    # ========================================================
    # OCR DETAILS
    # ========================================================

    with tab_ocr:

        st.subheader(
            "OCR details"
        )

        st.caption(
            "Use this only when something "
            "was missed or misread. "
            "The user-entered total remains "
            "the primary safety control."
        )

        st.text_area(
            "Raw OCR text",
            value=(
                st.session_state
                .raw_ocr
            ),
            height=500,
            disabled=True,
        )
